from openpyxl import load_workbook
#from mongoengine import connect
from pymongo import MongoClient
import sqlite3

client = MongoClient("mongodb://admin:admin@cluster0-shard-00-00-umdst.mongodb.net:27017,cluster0-shard-00-01-umdst.mongodb.net:27017,cluster0-shard-00-02-umdst.mongodb.net:27017/test?ssl=true&replicaSet=Cluster0-shard-0&authSource=admin&retryWrites=true")
db = client.BasePPPDigital

def importaBase():

    coll = db.Cotacao


    wb = load_workbook('ModelBase2.xlsx')
    sheet = wb['Base_Ativos']
    #print(sheet.cell(column=1, row=1).value)
    row= 2
    vet = []

    #print(sheet.cell(column=1, row=row).value)

    while row < 67560:

        doc= {
                'fonte': sheet.cell(column=1, row=row).value,
                'desc_pag':sheet.cell(column=2, row=row).value,
                'ticker':sheet.cell(column=3, row=row).value,
                'desc_ticker':sheet.cell(column=4, row=row).value,
                'mercado':sheet.cell(column=5, row=row).value.upper(),
                'tp_instrumento':sheet.cell(column=6, row=row).value
            }

        vet.append(doc)

        row = row +1

        if sheet.cell(column=1, row=row).value == None:
            break

    coll.insert_many(vet)

def consultaBase():

    con = sqlite3.connect('Base.db')
    c = con.cursor()
    a = 1
    for x in c.execute('SELECT * FROM Base_Ativos'):
        #print(x[1])
        print("Fonte:{}|  Mercado:{}| Tipo Ativo:{}".format(x[0],x[4],x[5]))

        a = a +1

def cotVisuPrin():
    coll = db.Cotacao

    vet = []

    for doc in coll.find({}, {"_id": 0, "desc_pag": 0, "desc_ticker": 0, "desc_pag": 0, "ticker": 0}):

        vet.append(doc)

    return vet



importaBase()
#consultaBase()

#a = cotVisuPrin()
#print(a)
